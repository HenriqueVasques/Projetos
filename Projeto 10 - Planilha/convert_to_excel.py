
import tkinter as tk
from customtkinter import filedialog
import os
import openpyxl
from docx import Document
import PyPDF2
import customtkinter as ctk

def format_cell_value(sheet, row_index, col_index, value):
    try:
        numeric_value = float(value)
        sheet.cell(row=row_index, column=col_index, value=numeric_value)
        sheet.cell(row=row_index, column=col_index).number_format = '0'  
    except ValueError:
        
        sheet.cell(row=row_index, column=col_index, value=value)


def add_borders(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style="thin", color="000000"),
                right=openpyxl.styles.Side(border_style="thin", color="000000"),
                top=openpyxl.styles.Side(border_style="thin", color="000000"),
                bottom=openpyxl.styles.Side(border_style="thin", color="000000")
            )

def txt_to_excel(lines, excel_filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row_index = 1
    for line in lines:
        values = line.strip().split()  
        if len(values) == 1:  
            values = line.strip().split('\t')
            if len(values) == 1:  
                values = line.strip().split('\t\t')
        for col_index, value in enumerate(values, start=1):
            format_cell_value(sheet, row_index, col_index, value)
        row_index += 1
    add_borders(sheet)
    workbook.save(excel_filename)

def docx_to_excel(paragraphs, excel_filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row_index = 1
    for paragraph in paragraphs:
        values = paragraph.text.strip().split()  
        if len(values) == 1:  
            values = paragraph.text.strip().split('\t')
            if len(values) == 1:  
                values = paragraph.text.strip().split('\t\t')
        for col_index, value in enumerate(values, start=1):
            format_cell_value(sheet, row_index, col_index, value)
        row_index += 1
    add_borders(sheet)
    workbook.save(excel_filename)

def pdf_to_excel(pdf_filename, excel_filename):
    text = ''
    with open(pdf_filename, 'rb') as pdf_file:
        reader = PyPDF2.PdfReader(pdf_file)
        for page in range(len(reader.pages)):
            text += reader.pages[page].extract_text()
    
    lines = text.split('\n')
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row_index = 1
    for line in lines:
        values = line.strip().split()  
        if len(values) == 1:  
            values = line.strip().split('\t')
            if len(values) == 1:  
                values = line.strip().split('\t\t')
        for col_index, value in enumerate(values, start=1):
            format_cell_value(sheet, row_index, col_index, value)
        row_index += 1
    add_borders(sheet)
    workbook.save(excel_filename)




def process_files(files, excel_filename):
    for file in files:
        filename, ext = os.path.splitext(file.name)
        if ext == '.txt':
            txt_to_excel(file.readlines(), excel_filename)
        elif ext == '.docx':
            doc = Document(file.name)
            docx_to_excel(doc.paragraphs, excel_filename)
        elif ext == '.pdf':
            pdf_to_excel(file.name, excel_filename)

# Procurar arquivos
def browse_files():
    # Solicita ao usuário para selecionar os arquivos
    files = filedialog.askopenfiles(filetypes=[("Text files", "*.txt"), ("Word files", "*.docx"), ("PDF files", "*.pdf")])
    
    if files:
        # Solicitar ao usuário para selecionar o diretório de salvamento
        save_filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        
        if save_filename:
            # Processar os arquivos e salvar o Excel no diretório selecionado
            process_files(files, save_filename)

# Janela Custom tkinker 

root = ctk.CTk()
root.title("Converter Arquivos para Excel")
root.geometry("500x300")


label = ctk.CTkLabel(root, text="Arraste e solte os arquivos aqui ou clique em 'Buscar Arquivos' para selecioná-los.")
label.pack(padx=10, pady=10)

button = ctk.CTkButton(root, text="Buscar Arquivos", command=browse_files)
button.pack(padx=10)
button.place(x = 180, y = 150)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

root.mainloop()

